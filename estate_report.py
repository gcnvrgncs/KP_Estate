import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
def generate_estate_report(data, filename="Estate_Report.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт о недвижимости"

    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = ["№ недвижимости", "Расположение", "Дата постройки", "Компания", "Контакты"]
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment

    for row_num, row_data in enumerate(data, 2):  # Начиная со 2 строки
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    for col_num, _ in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].auto_size = True

    workbook.save(filename)
    